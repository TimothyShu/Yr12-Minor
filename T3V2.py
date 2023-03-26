import openpyxl as pyxl
import FileManager
import csv
from pathlib import Path

#This creates the employee object

class Company:
    def __init__(self, CompName, compFileName, OvertimePay):
        self.OvertimePay = OvertimePay
        self.Name = CompName
        self.Employees = {}
        self.EmployeeIDs = []
        self.totalEmployees = len(self.Employees)
        self.FileName = compFileName
        self.wb, self.FilePath = FileManager.GetFile("T3V2Folder", self.FileName, GetPath=True)
        self.errorFile = Path("T3Dir") / f"{self.Name}log.csv"
        self.errors = []

    def AddEmployee(self, EmpID, Name, OvertimeHours, Pay):
        newEmployee = employee(EmpID, Name, OvertimeHours, Pay)
        if newEmployee.valid == False:
            print(f"\nThe employee is not valid\n")
            newEmployee.show()
        elif newEmployee.ID in self.EmployeeIDs:
            print(f"\nThe employee has a duplicate ID {newEmployee.ID}\n")
            newEmployee.show()
        else:
            self.Employees[str(EmpID)] = [Name, OvertimeHours, Pay]
            self.EmployeeIDs.append(newEmployee.ID)

    def FireEmployee(self, EmpID):
        try:
            del self.Employees[str(EmpID)]
            return True
        except KeyError:
            return False

    def readEmployees(self):
        ws1 = self.wb["Employees"]
        row = 2
        entry = [ws1.cell(row=row, column=n).value for n in range(1,5)]
        while entry[0]:
            self.AddEmployee(entry[0], entry[1], entry[2], entry[3])
            row += 1
            entry = [ws1.cell(row=row, column=n).value for n in range(1,5)]


    def updateEmployeeOvertime(self):
        WorkSheet = self.wb["Work"]
        row = 2
        entry = [WorkSheet.cell(row=row, column=n).value for n in range(1,3)]
        while entry[0]:
            #update the values of the employees
            if str(entry[0]) in self.Employees.keys():
                self.Employees[str(entry[0])][1] += entry[1]
            else:
                print(f"ID's invalid, errors have been movedinto {self.errorFile}")
                self.errors.append(f"invalid ID: {str(entry[0])} on row {row}")
            row += 1
            entry = [WorkSheet.cell(row=row, column=n).value for n in range(1,3)]

    def payEmployees(self):
        for key, item in self.Employees.items():
            self.Employees[key][2]  = item[2] + self.OvertimePay * item[1]
        

    def updateErrorSheet(self):
        try:
            FilePaths = open(self.errorFile, "x")
            writer = csv.writer(FilePaths, dialect = 'unix')
        except FileExistsError:
            FilePaths = open(self.errorFile, "w")
            writer = csv.writer(FilePaths, dialect = 'unix')

        for n in self.errors:
            print(n)
            writer.writerow([n])

        FilePaths.close()

    def updateEmployeeSheet(self):
        WorkSheet = self.wb["Employees"]
        row = 2
        for key, value in self.Employees.items():
            if str(WorkSheet.cell(row = row, column = 1).value) == key:
                WorkSheet.cell(row = row, column = 3).value = value[1]
                WorkSheet.cell(row=row, column = 4).value = value[2]
            else:
                pass
            row += 1
        self.wb.save(self.FilePath)

    def ResetEmployees(self):
        for key, item in self.Employees.items():
            self.Employees[key][1] = 0
        self.updateEmployeeSheet()

    def Iterate(self):
        self.ResetEmployees()
        self.readEmployees()
        self.updateEmployeeOvertime()
        self.updateErrorSheet()
        self.payEmployees()
        self.updateEmployeeSheet()
        self.wb.save(self.FilePath)
        pass
        

class employee:
    def __init__(self, EmpID, Name, OvertimeHours, Pay):
        self.valid = True
        self.ID = str(EmpID)
        self.Name = Name
        self.Overtime = OvertimeHours
        self.Pay = Pay
        
        if len(str(EmpID)) != 5:
            self.valid = False
            self.ID = "Invalid"

        try:
            EmpID = int(EmpID)
        except ValueError:
            pass

        if isinstance(EmpID, int) == False:
            self.valid = False
            self.ID = "Invalid"
        
    def show(self):
        print("Employee details are:\n")
        print(f"Is valid: {self.valid}")
        print(f"EmployeeID: {self.ID}")
        print(f"EmployeeName: {self.Name}")
        print(f"EmployeePay: {self.Pay}")
        print(f"EmployeeOvertime: {self.Overtime}")


inputs = {
    "Add" : ['a', 'add', 'append', 'create'],
    "Iterate" : ['i', 'iter', 'iterate', 'cycle', 'next cycle'],
    }

CurrentCompanies =[]


def IterateCompany():
    option = input("Please choose the Company You want to iterate over").lower()
    while option:
        for Company in CurrentCompanies:
            if Company.Name.lower() == option:
                validation = input(f"Are you certain you want to iterate over {Company.Name}?").lower()
                if validation.lower() == 'y':
                    print("Iterating through {Company.Name}")
                    Company.Iterate()
                break
        else:
            print("The Company does not exist")
        option = input("Please choose the Company You want to iterate over").lower()
        
                
def IterateAll():
    validation = input("Are you sure you want to iterate over all Companies?").lower()
    if validation == 'y':
        pass
    else:
        return
    for Company in CurrentCompanies:
        try:
            Company.Iterate()
            print(f"{Company.Name} has been iterated")
        except:
            print(f"The company {Company.Name} has an error")

def AddCompany():
    try:
        CompanyName = input("What is the company name: ")
        CompanyFiles = input("What is the company file name: ")
        CompanyPay = int(input("What is the hourly pay at said company: "))
    except ValueError:
        print("Entered wrong value for Pay")

    try:
        CurrentCompanies.append(Company(CompanyName, CompanyFiles, CompanyPay))
        print("Successfully added to Current Companies")
    except TypeError:
        print("Does CompanyFiles exist?")


def main():
    option = input("Select an option: ")
    while option:
        if option in inputs["Add"]:
            AddCompany()
        elif option in inputs["Iterate"]:
            choice = input("Do you want to iterate over a Company or all Companies?").lower()
            if choice == "all":
                IterateAll()
            elif choice == "one":
                IterateCompany()
        option = input("Select an option: ")


if __name__ == "__main__":
    CurrentCompanies.append(Company("St Joseph's College", "St Joseph's College Staff.xlsx", 10))
    CurrentCompanies.append(Company("King's College", "King's College Staff.xlsx", 20))
    CurrentCompanies.append(Company("Scot's College", "Scot's College Staff.xlsx", 15))
    main()
