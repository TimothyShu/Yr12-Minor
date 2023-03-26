import openpyxl as pyxl
from random import randint
import FileManager

wb, filepath = FileManager.GetFile("Root", "JoeysStaff.xlsx", GetPath=True)
ws = wb["Employees"]
ws1 = wb["Work"]
letters = "abcdefghijklmnopqrstuvwxyz"
employeeIDs = []

def createID():
    ID = "".join([str(randint(0,9)) for n in range(0,5)])
    if ID not in employeeIDs:
        employeeIDs.append(ID)
    else:
        ID = createID()
    return ID

def createName():
    name = "".join([letters[randint(0,25)] for n in range(0, randint(5,15))])
    return name

def createTime():
    time = 0.25*randint(0,20)
    return time


def clear(ws):
    row =2
    for n in range(0, 500):
        #ID
        ws.cell(row=row, column=1).value = None
        #Name
        ws.cell(row=row, column=2).value = None
        #Overtime
        ws.cell(row=row, column=3).value = None
        row += 1

def createEmployees(number = 10):
    row = 2
    for n in range(0, number):
        #ID
        ws.cell(row=row, column=1).value = createID()
        #Name
        ws.cell(row=row, column=2).value = createName()
        #Overtime
        ws.cell(row=row, column=3).value = createTime()
        
        row =row + 1

        
def readEmployees():
    row = 2
    entry = [ws.cell(row=row, column=n).value for n in range(1,4)]
    employees = []
    while entry[0]:
        employees.append([entry[0], entry[1], entry[2]])
        row += 1
        entry = [ws.cell(row=row, column=n).value for n in range(1,4)]
    return employees

        
def createworkfiles(number = 150):
    employees = readEmployees()
    employees.append([194769, "Tim", 1.5])
    row = 2
    for n in range(0, number):
        #ID
        ws1.cell(row=row, column=1).value = employees[randint(0, len(employees)-1)][0]
        #Overtime
        ws1.cell(row=row, column=2).value = createTime()
        row =row + 1
    pass

def save():
    wb.save(filepath)

#clear(ws1)
#createEmployees()
#createworkfiles()

wb.save(filepath)
