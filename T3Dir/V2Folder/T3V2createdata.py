import openpyxl as pyxl
from random import randint
import FileManager
from pathlib import Path

letters = "abcdefghijklmnopqrstuvwxyz"

def loadFile(FileName):
    global wb, filepath, ws, ws1, employeeIDs
    wb, filepath = FileManager.GetFile("Root", FileName, GetPath=True)
    ws = wb["Employees"]
    ws1 = wb["Work"]
    employeeIDs = []

def createID():
    ID = "".join([str(randint(0,9)) for n in range(0,5)])
    if ID not in employeeIDs:
        employeeIDs.append(ID)
    else:
        ID = createID()
    return ID

def createName():
    name = "".join([letters[randint(0,25)] for n in range(0, randint(5,15))])
    return name

def createTime():
    time = 0.25*randint(0,20)
    return time


def clear(ws):
    row =2
    for n in range(0, 500):
        #ID
        ws.cell(row=row, column=1).value = None
        #Name
        ws.cell(row=row, column=2).value = None
        #Overtime
        ws.cell(row=row, column=3).value = None
        row += 1

def createEmployees(wb, number = 10):
    ws = wb["Employees"]
    row = 2
    for n in range(0, number):
        #ID
        ws.cell(row=row, column=1).value = createID()
        #Name
        ws.cell(row=row, column=2).value = createName()
        #Overtime
        ws.cell(row=row, column=3).value = createTime()
        #payment
        ws.cell(row=row, column =4).value = 0
        
        row =row + 1

        
def readEmployees(wb):
    ws = wb["Employees"]
    row = 2
    entry = [ws.cell(row=row, column=n).value for n in range(1,5)]
    employees = []
    while entry[0]:
        employees.append([entry[0], entry[1], entry[2]])
        row += 1
        entry = [ws.cell(row=row, column=n).value for n in range(1,5)]
    return employees

        
def createworkfiles(wb, number = 150):
    ws1 = wb["Work"]
    employees = readEmployees(wb)
    employees.append([194769, "Tim", 1.5])
    row = 2
    for n in range(0, number):
        #ID
        ws1.cell(row=row, column=1).value = employees[randint(0, len(employees)-1)][0]
        #Overtime
        ws1.cell(row=row, column=2).value = createTime()
        row =row + 1
    pass

def createHeader(wb):
    ws1 = wb["Employees"]
    ws1.cell(row=1, column=1).value = "EmpID"
    ws1.cell(row=1, column=2).value = "Name"
    ws1.cell(row=1, column=3).value = "OverTime"
    ws1.cell(row=1, column=4).value = "Pay"
    ws2 = wb["Work"]
    ws2.cell(row=1, column=1).value = "EmpID"
    ws2.cell(row=1, column=2).value = "OverTime"

def createCompanyFile(CompanyName):
    current_path = Path(FileManager.GetDirPath("Root"))
    file = f"{CompanyName} Staff.xlsx"
    wb = pyxl.Workbook()
    EmployeeSheet = wb.create_sheet("Employees", 0)
    WorkSheet = wb.create_sheet("Work", 1)
    createHeader(wb)
    wb.save(file)
    loadFile(file)

def createCompanyData(n_employees = 10, n_workfiles = 150):
    readEmployees(wb)
    createEmployees(wb, n_employees)
    save()
    createworkfiles(wb, n_workfiles)
    save()

def CreateTestCompany(CompanyName, employees =10, workfiles = 150):
    createCompanyFile(CompanyName)
    createCompanyData(employees, workfiles)

def save():
    wb.save(filepath)


