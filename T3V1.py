import openpyxl as pyxl
import FileManager
import csv
from pathlib import Path

#This creates the employee object

class Company:
    def __init__(self, CompName, compFileName):
        self.Name = CompName
        self.Employees = {}
        self.EmployeeIDs = []
        self.totalEmployees = len(self.Employees)
        self.FileName = compFileName
        self.wb, self.FilePath = FileManager.GetFile("T3Dir", self.FileName, GetPath=True)
        self.errorFile = Path("T3Dir") / f"{self.Name}log.csv"
        self.errors = []

    def AddEmployee(self, EmpID, Name, OvertimeHours):
        newEmployee = employee(EmpID, Name, OvertimeHours)
        if newEmployee.valid == False:
            print(f"\nThe employee is not valid\n")
            newEmployee.show()
        elif newEmployee.ID in self.EmployeeIDs:
            print(f"\nThe employee has a duplicate ID {newEmployee.ID}\n")
            newEmployee.show()
        else:
            self.Employees[str(EmpID)] = [Name, OvertimeHours]
            self.EmployeeIDs.append(newEmployee.ID)

    def FireEmployee(self, EmpID):
        try:
            del self.Employees[str(EmpID)]
            return True
        except KeyError:
            return False

    def readEmployees(self):
        ws1 = self.wb["Employees"]
        row = 2
        entry = [ws1.cell(row=row, column=n).value for n in range(1,4)]
        while entry[0]:
            self.AddEmployee(entry[0], entry[1], entry[2])
            row += 1
            entry = [ws1.cell(row=row, column=n).value for n in range(1,4)]


    def updateEmployeeOvertime(self):
        WorkSheet = self.wb["Work"]
        row = 2
        entry = [WorkSheet.cell(row=row, column=n).value for n in range(1,3)]
        while entry[0]:
            #update the values of the employees
            if str(entry[0]) in self.Employees.keys():
                self.Employees[str(entry[0])][1] += entry[1]
            else:
                print(f"ID's invalid, errors have been movedinto {self.errorFile}")
                self.errors.append(f"invalid ID: {str(entry[0])} on row {row}")
            row += 1
            entry = [WorkSheet.cell(row=row, column=n).value for n in range(1,3)]
        self.updateErrorSheet()
        self.updateEmployeeSheet()
        self.wb.save(self.FilePath)

    def updateErrorSheet(self):
        try:
            FilePaths = open(self.errorFile, "x")
            writer = csv.writer(FilePaths, dialect = 'unix')
        except FileExistsError:
            FilePaths = open(self.errorFile, "w")
            writer = csv.writer(FilePaths, dialect = 'unix')

        for n in self.errors:
            print(n)
            writer.writerow([n])

        FilePaths.close()

    def updateEmployeeSheet(self):
        WorkSheet = self.wb["Employees"]
        row = 2
        for key, value in self.Employees.items():
            if str(WorkSheet.cell(row = row, column = 1).value) == key:
                WorkSheet.cell(row = row, column = 3).value = value[1]
            else:
                pass
            row += 1
        pass

    def ResetEmployees(self):
        for key, item in self.Employees.items():
            self.Employees[key][1] = 0
        self.updateEmployeeSheet()
        self.wb.save(self.FilePath)
        

class employee:
    def __init__(self, EmpID, Name, OvertimeHours):
        self.valid = True
        self.ID = str(EmpID)
        self.Name = Name
        self.Overtime = OvertimeHours
        
        if len(str(EmpID)) != 5:
            self.valid = False
            self.ID = "Invalid"

        try:
            EmpID = int(EmpID)
        except ValueError:
            pass

        if isinstance(EmpID, int) == False:
            self.valid = False
            self.ID = "Invalid"
        
    def show(self):
        print("Employee details are:\n")
        print(f"Is valid: {self.valid}")
        print(f"EmployeeID: {self.ID}")
        print(f"EmployeeName: {self.Name}")
        print(f"EmployeeOvertime: {self.Overtime}")



if __name__ == "__main__":
    Joeys = Company("Joeys", "JoeysStaff.xlsx")
    Joeys.readEmployees()
    Joeys.updateEmployeeOvertime()
